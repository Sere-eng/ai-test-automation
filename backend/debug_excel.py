#!/usr/bin/env python3
"""
Debug script per analizzare il file Excel e capire come migliorare il parser.
"""

import sys
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent))

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Color
except ImportError:
    print("❌ openpyxl non installato. Installa con: pip install openpyxl")
    sys.exit(1)


def analyze_excel_structure(excel_file: Path):
    """Analizza la struttura del file Excel."""
    
    if not excel_file.exists():
        print(f"❌ File non trovato: {excel_file}")
        return
    
    print("=" * 80)
    print(f"ANALISI STRUTTURA EXCEL: {excel_file.name}")
    print("=" * 80)
    print()
    
    # Carica il workbook
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Header
    print("📋 HEADER (Riga 1):")
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    for i, cell in enumerate(header_row):
        if cell:
            print(f"  Colonna {i}: {cell}")
    print()
    
    # Analizza le prime 30 righe con dettagli
    print("=" * 80)
    print("📊 ANALISI RIGHE (Prime 30 righe con contenuto)")
    print("=" * 80)
    print()
    
    rows_analyzed = 0
    current_color = None
    color_group = 1
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if rows_analyzed >= 30:
            break
        
        # Ottieni i valori
        row_values = [cell.value for cell in row]
        
        # Verifica se vuota
        is_empty = all(v is None or str(v).strip() == '' for v in row_values)
        if is_empty:
            continue
        
        rows_analyzed += 1
        
        # Ottieni il colore di sfondo delle prime 3 celle
        colors = []
        for i in range(min(3, len(row))):
            cell = row[i]
            if cell and cell.fill:
                fill = cell.fill
                if fill.start_color and hasattr(fill.start_color, 'rgb') and fill.start_color.rgb:
                    colors.append(str(fill.start_color.rgb))
                elif fill.fgColor and hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                    colors.append(str(fill.fgColor.rgb))
                else:
                    colors.append('NONE')
            else:
                colors.append('NONE')
        
        # Determina il colore principale
        row_color = colors[0] if colors else 'NONE'
        
        # Rileva cambio colore
        if current_color != row_color:
            if current_color is not None:
                print()
                print(f"{'='*80}")
                print()
            current_color = row_color
            color_group += 1
            print(f"🎨 GRUPPO COLORE #{color_group - 1} - Colore: {row_color}")
            print(f"   Colori celle [0,1,2]: {colors}")
            print()
        
        print(f"Riga {row_idx}:")
        
        # Mostra solo le prime 5 colonne con contenuto
        for i, value in enumerate(row_values[:5]):
            if value and str(value).strip():
                value_str = str(value).replace('\n', ' ')[:80]
                if len(str(value)) > 80:
                    value_str += "..."
                print(f"  Col {i} ({header_row[i] if i < len(header_row) else 'N/A'}): {value_str}")
        print()
    
    print("=" * 80)
    print(f"✅ Analizzate {rows_analyzed} righe con contenuto")
    print("=" * 80)


def compare_with_parser():
    """Confronta l'analisi manuale con il risultato del parser."""
    from agent.document_parser import parse_test_document
    
    excel_file = Path(__file__).parent.parent / "Casi di test di CPOE-T validi il 03.03.2026 (1).csv"
    
    # Cerca file xlsx nella cartella corrente
    xlsx_files = list(Path(__file__).parent.parent.glob("*.xlsx"))
    if xlsx_files:
        excel_file = xlsx_files[0]
    
    if not excel_file.exists():
        print(f"\n❌ File Excel non trovato. Cerca: {excel_file}")
        return
    
    print("\n\n")
    print("=" * 80)
    print("CONFRONTO CON PARSER")
    print("=" * 80)
    print()
    
    try:
        result = parse_test_document(str(excel_file))
        
        print(f"📄 Formato: {result.get('format')}")
        print(f"📊 Test cases estratti: {result.get('test_cases_count')}")
        print()
        
        # Mostra i primi 3 test case
        test_cases = result.get('test_cases', [])
        for i, tc in enumerate(test_cases[:3], 1):
            print(f"--- TEST CASE #{i} ---")
            print(f"Riga Excel: {tc.get('row_number')}")
            print()
            
            if tc.get('objective'):
                print(f"OBIETTIVO: {tc['objective'][:100]}...")
                print()
            
            if tc.get('prerequisites'):
                lines = tc['prerequisites'].split('\n')
                print(f"PREREQUISITI ({len(lines)} righe):")
                for line in lines[:5]:
                    print(f"  {line[:80]}")
                if len(lines) > 5:
                    print(f"  ... (+{len(lines)-5} righe)")
                print()
            
            print()
        
    except Exception as e:
        print(f"❌ Errore nel parser: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # Cerca il file Excel
    excel_file = Path(__file__).parent.parent / "Casi di test di CPOE-T validi il 03.03.2026 (1).csv"
    
    # Cerca file xlsx nella cartella corrente o parent
    xlsx_files = list(Path(__file__).parent.parent.glob("*.xlsx"))
    xlsx_files.extend(list(Path(__file__).parent.glob("*.xlsx")))
    
    if xlsx_files:
        excel_file = xlsx_files[0]
        print(f"📁 File trovato: {excel_file}")
    elif excel_file.exists():
        print(f"📁 File trovato (CSV): {excel_file}")
    else:
        print("❌ Nessun file Excel/CSV trovato.")
        print("   Cerca in:")
        print(f"   - {Path(__file__).parent.parent}")
        print(f"   - {Path(__file__).parent}")
        sys.exit(1)
    
    print()
    
    # Esegui analisi
    if excel_file.suffix.lower() == '.xlsx':
        analyze_excel_structure(excel_file)
        compare_with_parser()
    else:
        print("⚠️  File CSV trovato - usa formato XLSX per analisi colori")
        compare_with_parser()
