# backend/agent/document_parser.py
"""
Parser per documenti di test (Word, HTML, XLSX).
Estrae sezioni strutturate da documenti JIRA esportati o test case manuali.
"""

import re
from pathlib import Path
from typing import Dict, Optional, List
from bs4 import BeautifulSoup
import chardet


class TestDocumentParser:
    """Parser per documenti di test case in vari formati."""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.content = None
        self.soup = None
        
    def parse(self) -> Dict[str, str]:
        """
        Estrae le sezioni principali dal documento.
        
        Returns:
            Dict con chiavi: title, initial_conditions, test_steps, expected_results, raw_html
            Per Excel/CSV: ritorna anche test_cases (lista di casi di test)
        """
        if not self.file_path.exists():
            raise FileNotFoundError(f"File non trovato: {self.file_path}")
        
        # Determina il tipo di file dall'estensione
        file_ext = self.file_path.suffix.lower()
        
        # Per file Excel/CSV, usa parsing strutturato
        if file_ext in ['.xlsx', '.xls', '.csv']:
            return self._parse_spreadsheet()
        
        # Per altri file, leggi il contenuto con encoding detection
        raw_bytes = self.file_path.read_bytes()
        detected = chardet.detect(raw_bytes)
        encoding = detected['encoding'] or 'utf-8'
        
        try:
            self.content = raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            # Fallback a latin-1 se UTF-8 fallisce
            self.content = raw_bytes.decode('latin-1')
        
        # Determina il tipo di file
        if self._is_html():
            return self._parse_html()
        elif file_ext == '.docx':
            return self._parse_docx()
        else:
            # File .doc che in realtà è HTML (come quelli esportati da JIRA)
            return self._parse_html()
    
    def _is_html(self) -> bool:
        """Verifica se il contenuto è HTML."""
        return bool(re.search(r'<html|<HTML|<!DOCTYPE', self.content[:500]))
    
    def _parse_html(self) -> Dict[str, str]:
        """
        Estrae informazioni da HTML (tipicamente export JIRA).
        
        Cerca:
        - Titolo (h3.formtitle o primo h3)
        - Condizioni Iniziali (td con "Condizioni Iniziali")
        - Passi del Test (td con "Passi del Test")
        - Condizioni Finali (td con "Condizioni Finali")
        """
        self.soup = BeautifulSoup(self.content, 'html.parser')
        
        result = {
            'title': self._extract_title(),
            'initial_conditions': self._extract_section('Condizioni Iniziali', 'Dati di input'),
            'test_steps': self._extract_section('Passi del Test', 'Modalità di esecuzione'),
            'expected_results': self._extract_section('Condizioni Finali', 'Risultati attesi'),
            'raw_html': self.content,
            'format': 'html'
        }
        
        return result
    
    def _parse_docx(self) -> Dict[str, str]:
        """
        Estrae informazioni da file Word .docx.
        Richiede python-docx.
        """
        try:
            from docx import Document
        except ImportError:
            raise ImportError(
                "python-docx non installato. Installa con: pip install python-docx"
            )
        
        doc = Document(self.file_path)
        
        # Estrai tutto il testo
        full_text = '\n'.join([para.text for para in doc.paragraphs])
        
        # Pattern per trovare sezioni
        sections = {
            'title': self._extract_docx_section(full_text, r'^(.+?)$', first_line=True),
            'initial_conditions': self._extract_docx_section(
                full_text, 
                r'(?:Condizioni Iniziali|Prerequisiti)[:\s]+(.*?)(?=Condizioni Finali|Passi del Test|Modalità|$)',
                re.DOTALL | re.IGNORECASE
            ),
            'test_steps': self._extract_docx_section(
                full_text,
                r'(?:Passi del Test|Modalità di esecuzione)[:\s]+(.*?)(?=Condizioni Finali|Risultati attesi|Dati di input|$)',
                re.DOTALL | re.IGNORECASE
            ),
            'expected_results': self._extract_docx_section(
                full_text,
                r'(?:Condizioni Finali|Risultati attesi)[:\s]+(.*?)(?=Dati di input|$)',
                re.DOTALL | re.IGNORECASE
            ),
            'raw_text': full_text,
            'format': 'docx'
        }
        
        return sections
    
    def _parse_spreadsheet(self) -> Dict[str, str]:
        """
        Estrae informazioni da file Excel (.xlsx, .xls) o CSV.
        
        Formato atteso: 
        - Colonne: OBIETTIVO, PREREQUISITI, DATI_INPUT, DESCRIZIONE, RISULTATI_ATTESI
        - Ogni test case si estende su più righe con lo stesso colore di sfondo
        - Righe con colore alternato (bianco/blu) indicano test diversi
        
        Returns:
            Dict con test_cases (lista di dict con i test) e metadati
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError(
                "pandas non installato. Installa con: pip install pandas openpyxl xlrd"
            )
        
        file_ext = self.file_path.suffix.lower()
        
        # Per file Excel, usa openpyxl per accedere ai colori
        if file_ext in ['.xlsx']:
            return self._parse_excel_with_colors()
        elif file_ext == '.xls':
            # .xls vecchio formato, usa pandas con engine xlrd
            try:
                df = pd.read_excel(self.file_path, engine="xlrd")
            except ImportError as exc:
                raise ImportError(
                    "Impossibile leggere file .xls: il pacchetto 'xlrd' non è installato. "
                    "Installa con: pip install xlrd"
                ) from exc
            return self._parse_dataframe_standard(df, file_ext)
        elif file_ext == '.csv':
            # CSV - parsing standard
            # Prova diversi encoding e separatori comuni
            for encoding in ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']:
                for sep in [';', ',', '\t']:
                    try:
                        df = pd.read_csv(self.file_path, encoding=encoding, sep=sep)
                        # Verifica che abbia almeno 2 colonne
                        if len(df.columns) >= 2:
                            return self._parse_dataframe_standard(df, file_ext)
                    except (UnicodeDecodeError, pd.errors.ParserError):
                        continue
            # Se tutti i tentativi falliscono, usa il default
            df = pd.read_csv(self.file_path, encoding='utf-8', sep=';')
            return self._parse_dataframe_standard(df, file_ext)
        else:
            raise ValueError(f"Formato file non supportato: {file_ext}")
    
    def _parse_excel_with_colors(self) -> Dict[str, str]:
        """
        Parse Excel usando openpyxl per accedere ai colori delle celle.
        Raggruppa righe con lo stesso colore di sfondo.
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Color
        except ImportError:
            raise ImportError(
                "openpyxl non installato. Installa con: pip install openpyxl"
            )
        
        # Carica il workbook
        wb = load_workbook(self.file_path)
        ws = wb.active
        
        # Leggi l'header (prima riga)
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        # Identifica le colonne
        col_indices = self._identify_columns(header_row)
        
        print(f"📊 Colonne identificate: {col_indices}")
        
        # Raggruppa righe per colore
        test_cases = []
        current_test = None
        current_color = None
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # Ottieni il colore di sfondo - MIGLIORATO per guardare tutte le celle importanti
            row_color = self._get_row_background_color(row, col_indices)
            
            # Estrai i valori delle celle
            row_values = [cell.value for cell in row]
            
            # Verifica se la riga è completamente vuota
            is_empty = all(v is None or str(v).strip() == '' for v in row_values)
            
            if is_empty:
                continue
            
            # Determina se questa riga appartiene a un nuovo test case
            # LOGICA: cambio di colore = nuovo test
            # Ma consideriamo solo cambi SIGNIFICATIVI (non da None a qualcosa)
            is_new_test = False
            
            if current_color is None:
                # Prima riga
                is_new_test = True
            elif row_color != current_color and row_color is not None:
                # Colore cambiato
                is_new_test = True
            
            if is_new_test:
                # Salva il test precedente se esiste e ha contenuto
                if current_test and self._test_has_content(current_test):
                    test_cases.append(current_test)
                    print(f"✓ Test case #{len(test_cases)} salvato")
                
                # Inizia un nuovo test case
                current_test = {
                    'row_number': row_idx,
                    'objective': '',
                    'prerequisites': '',
                    'input_data': '',
                    'description': '',
                    'expected_results': '',
                }
                current_color = row_color
                print(f"→ Nuovo test case iniziato alla riga {row_idx} (colore: {row_color})")
            else:
                # Riga di continuazione
                print(f"  + Riga {row_idx} aggiunta (colore: {row_color})")
            
            # Aggiungi i valori al test corrente
            if current_test:
                for field, col_idx in col_indices.items():
                    if col_idx is not None and col_idx < len(row_values):
                        value = self._clean_cell_value(row_values[col_idx])
                        if value:
                            # Concatena con newline se già presente
                            if current_test[field]:
                                current_test[field] += '\n' + value
                            else:
                                current_test[field] = value
        
        # Aggiungi l'ultimo test case se ha contenuto
        if current_test and self._test_has_content(current_test):
            test_cases.append(current_test)
            print(f"✓ Test case #{len(test_cases)} salvato (ultimo)")
        
        # Rimuovi il campo 'color' dai risultati (era solo per debugging)
        for tc in test_cases:
            tc.pop('color', None)
        
        print(f"\n📝 Totale test case estratti: {len(test_cases)}")
        
        # Crea un titolo dal nome del file
        title = self.file_path.stem.replace('_', ' ').replace('-', ' ')
        
        return {
            'title': title,
            'format': 'xlsx',
            'test_cases_count': len(test_cases),
            'test_cases': test_cases,
            'columns_found': {k: v for k, v in col_indices.items()}
        }
    
    def _get_row_background_color(self, row, col_indices) -> Optional[str]:
        """
        Ottiene il colore di sfondo della riga controllando le celle significative.
        Ignora celle vuote e considera solo celle con contenuto.
        """
        # Controlla TUTTE le celle con contenuto, non solo la prima
        colors_found = []
        
        for cell in row:
            if cell and cell.value is not None and str(cell.value).strip() != '':
                # Questa cella ha contenuto, controlla il colore
                if cell.fill:
                    fill = cell.fill
                    if fill.start_color and hasattr(fill.start_color, 'rgb') and fill.start_color.rgb:
                        color = str(fill.start_color.rgb)
                        if color not in ['00000000', 'FFFFFFFF']:  # Ignora bianco/trasparente
                            colors_found.append(color)
                    elif fill.fgColor and hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                        color = str(fill.fgColor.rgb)
                        if color not in ['00000000', 'FFFFFFFF']:
                            colors_found.append(color)
        
        # Se abbiamo trovato colori significativi, usa il più comune
        if colors_found:
            # Conta le occorrenze e restituisci il più frequente
            from collections import Counter
            most_common = Counter(colors_found).most_common(1)[0][0]
            return most_common
        
        # Se non ci sono colori significativi, controlla la prima cella comunque
        first_cell = row[0] if row else None
        if first_cell and first_cell.fill:
            fill = first_cell.fill
            if fill.start_color and hasattr(fill.start_color, 'rgb') and fill.start_color.rgb:
                return str(fill.start_color.rgb)
            elif fill.fgColor and hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                return str(fill.fgColor.rgb)
        
        return '00000000'  # Default bianco/trasparente
    
    def _is_color_significantly_different(self, color1, color2) -> bool:
        """
        Verifica se due colori sono significativamente diversi.
        Entrambi None = non diversi
        Uno None e uno set = diversi
        """
        if color1 is None and color2 is None:
            return False
        if color1 is None or color2 is None:
            return True
        return color1 != color2
    
    def _test_has_content(self, test_case: Dict) -> bool:
        """
        Verifica se un test case ha contenuto utile.
        """
        # Almeno uno dei campi principali deve avere contenuto
        return bool(
            test_case.get('objective') or 
            test_case.get('description') or 
            test_case.get('expected_results')
        )
    
    def _parse_dataframe_standard(self, df, file_ext: str) -> Dict[str, str]:
        """
        Parse standard di dataframe per CSV o XLS senza informazioni di colore.
        Ogni riga non vuota è considerata un test case separato.
        """
        import pandas as pd
        
        # Normalizza i nomi delle colonne
        column_map = {}
        for col in df.columns:
            col_clean = str(col).strip().upper()
            column_map[col] = col_clean
        
        # Identifica le colonne rilevanti
        col_indices = {}
        for orig_col, clean_col in column_map.items():
            if 'OBIETTIVO' in clean_col or 'OBJECTIVE' in clean_col or 'FUNZIONE' in clean_col:
                col_indices['objective'] = orig_col
            elif 'PREREQUISIT' in clean_col or 'CONDIZIONI INIZIALI' in clean_col:
                col_indices['prerequisites'] = orig_col
            elif 'DATI' in clean_col and 'INPUT' in clean_col:
                col_indices['input_data'] = orig_col
            elif 'DESCRIZIONE' in clean_col or 'MODALIT' in clean_col or 'PASSI' in clean_col:
                col_indices['description'] = orig_col
            elif 'RISULTAT' in clean_col and 'ATTESI' in clean_col or 'EXPECTED' in clean_col or 'CONDIZIONI FINALI' in clean_col:
                col_indices['expected_results'] = orig_col
        
        # Estrai i casi di test
        test_cases = []
        for idx, row in df.iterrows():
            # Salta righe vuote
            if row.isna().all() or all(str(v).strip() == '' for v in row if pd.notna(v)):
                continue
            
            test_case = {
                'row_number': int(idx) + 2,  # +2 perché idx è 0-based e prima riga è header
                'objective': self._clean_cell_value(row.get(col_indices.get('objective'))) if 'objective' in col_indices else '',
                'prerequisites': self._clean_cell_value(row.get(col_indices.get('prerequisites'))) if 'prerequisites' in col_indices else '',
                'input_data': self._clean_cell_value(row.get(col_indices.get('input_data'))) if 'input_data' in col_indices else '',
                'description': self._clean_cell_value(row.get(col_indices.get('description'))) if 'description' in col_indices else '',
                'expected_results': self._clean_cell_value(row.get(col_indices.get('expected_results'))) if 'expected_results' in col_indices else '',
            }
            
            # Aggiungi solo se c'è almeno un campo non vuoto
            if any(test_case[k] for k in ['objective', 'description', 'expected_results']):
                test_cases.append(test_case)
        
        # Crea un titolo dal nome del file
        title = self.file_path.stem.replace('_', ' ').replace('-', ' ')
        
        return {
            'title': title,
            'format': file_ext.lstrip('.'),
            'test_cases_count': len(test_cases),
            'test_cases': test_cases,
            'columns_found': col_indices
        }
    
    def _identify_columns(self, header_row) -> Dict[str, int]:
        """
        Identifica gli indici delle colonne dall'header.
        
        Returns:
            Dict con field_name -> column_index (0-based)
        """
        col_indices = {
            'objective': None,
            'prerequisites': None,
            'input_data': None,
            'description': None,
            'expected_results': None
        }
        
        for idx, col_name in enumerate(header_row):
            if col_name is None:
                continue
            
            col_clean = str(col_name).strip().upper()
            
            if 'OBIETTIVO' in col_clean or 'OBJECTIVE' in col_clean or 'FUNZIONE' in col_clean:
                col_indices['objective'] = idx
            elif 'PREREQUISIT' in col_clean or 'CONDIZIONI INIZIALI' in col_clean:
                col_indices['prerequisites'] = idx
            elif 'DATI' in col_clean and 'INPUT' in col_clean:
                col_indices['input_data'] = idx
            elif 'DESCRIZIONE' in col_clean or 'MODALIT' in col_clean or 'PASSI' in col_clean:
                col_indices['description'] = idx
            elif ('RISULTAT' in col_clean and 'ATTESI' in col_clean) or 'EXPECTED' in col_clean or 'CONDIZIONI FINALI' in col_clean:
                col_indices['expected_results'] = idx
        
        return col_indices
    
    def _clean_cell_value(self, value) -> str:
        """
        Pulisce il valore di una cella Excel/CSV.
        Gestisce NaN, None, e formattazione.
        """
        import pandas as pd
        
        if pd.isna(value) or value is None:
            return ''
        
        # Converti a stringa
        text = str(value).strip()
        
        # Rimuovi caratteri speciali comuni nei CSV
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        
        # Rimuovi spazi multipli
        text = re.sub(r' +', ' ', text)
        
        # Rimuovi newline multiple
        text = re.sub(r'\n\s*\n+', '\n\n', text)
        
        # Gestisci il carattere ^ usato spesso per indicare prerequisiti concatenati
        # Sostituiscilo con un bullet point per migliore leggibilità
        if text.startswith('^'):
            text = text.replace('^', '• ', 1)
            text = text.replace('^', '\n• ')
        
        return text.strip()
    
    def _extract_title(self) -> str:
        """Estrae il titolo dal documento HTML."""
        # Prova h3.formtitle
        title_elem = self.soup.find('h3', class_='formtitle')
        if title_elem:
            # Rimuovi il ticket ID tipo [HC40DIAGOL-7155]
            text = title_elem.get_text(strip=True)
            text = re.sub(r'\[.*?\]\s*', '', text)
            return text
        
        # Fallback: primo h3
        h3 = self.soup.find('h3')
        if h3:
            text = h3.get_text(strip=True)
            text = re.sub(r'\[.*?\]\s*', '', text)
            return text
        
        # Fallback: title tag
        title = self.soup.find('title')
        if title:
            text = title.get_text(strip=True)
            text = re.sub(r'\[.*?\]\s*', '', text)
            return text
        
        return "Untitled"
    
    def _extract_section(self, *labels: str) -> str:
        """
        Estrae una sezione dal documento HTML cercando label specifiche.
        
        Args:
            *labels: Una o più label da cercare (es. "Condizioni Iniziali", "Prerequisiti")
        
        Returns:
            Testo della sezione pulito
        """
        for label in labels:
            label_lower = label.lower().strip()

            # Passo 1: versione "rigida" (bgcolor="#f0f0f0") per mantenere compatibilità
            cells = self.soup.find_all('td', bgcolor="#f0f0f0")
            for cell in cells:
                # Alcuni export hanno la label in <b> mentre altre varianti usano <span>/<strong>
                text = ""
                b_tag = cell.find('b')
                if b_tag:
                    text = b_tag.get_text(strip=True)
                else:
                    # fallback locale: prendi tutto il testo della cella se non c'è <b>
                    text = cell.get_text(strip=True)

                if text and label_lower in text.lower():
                    tr = cell.parent
                    tds = tr.find_all('td')
                    if len(tds) >= 2:
                        content_cell = tds[1]
                        return self._clean_html_text(content_cell)

            # Passo 2: fallback "tollerante" (non dipende dal bgcolor)
            # Cerca tag che contengono la label (es. <b>Condizioni Iniziali:</span>)
            for tag in self.soup.find_all(['b', 'strong', 'span']):
                try:
                    tag_text = tag.get_text(strip=True)
                except Exception:
                    tag_text = ""
                if not tag_text:
                    continue
                if label_lower not in tag_text.lower():
                    continue

                td = tag.find_parent('td')
                if not td:
                    continue

                tr = td.parent
                tds = tr.find_all('td')
                if len(tds) >= 2:
                    content_cell = tds[1]
                    return self._clean_html_text(content_cell)

        return ""
    
    def _clean_html_text(self, element) -> str:
        """
        Pulisce il testo HTML rimuovendo tag ma preservando struttura liste.
        """
        if not element:
            return ""
        
        # Sostituisci <li> con bullet point
        for li in element.find_all('li'):
            li.insert(0, '• ')
        
        # Sostituisci <p> con newline
        for p in element.find_all('p'):
            p.append('\n')
        
        # Estrai testo
        text = element.get_text(separator='\n', strip=True)
        
        # Pulisci spazi multipli
        text = re.sub(r'\n\s*\n+', '\n\n', text)
        text = re.sub(r' +', ' ', text)
        
        return text.strip()
    
    def _extract_docx_section(self, text: str, pattern: str, flags=0, first_line=False) -> str:
        """Helper per estrarre sezioni da testo .docx con regex."""
        if first_line:
            lines = text.split('\n')
            return lines[0] if lines else ""
        
        match = re.search(pattern, text, flags)
        if match:
            return match.group(1).strip() if match.lastindex else match.group(0).strip()
        return ""
    
    def extract_scenarios(self) -> list[Dict[str, str]]:
        """
        Estrae gli scenari individuali dal documento.
        Ogni scenario contiene: id, name, steps, expected_results.
        
        Returns:
            Lista di dict con scenari
        """
        data = self.parse()
        test_steps = data.get("test_steps", "")
        expected_results = data.get("expected_results", "")
        
        # Identifica scenari (Nel primo scenario, Nel secondo scenario, etc.)
        scenario_pattern = r'Nel (primo|secondo|terzo|quarto|quinto|sesto|settimo|ottavo|nono|decimo) scenario:'
        step_parts = re.split(scenario_pattern, test_steps, flags=re.IGNORECASE)
        result_parts = re.split(scenario_pattern, expected_results, flags=re.IGNORECASE)
        
        scenarios = []
        scenario_map = {
            "primo": 1, "secondo": 2, "terzo": 3, "quarto": 4, "quinto": 5,
            "sesto": 6, "settimo": 7, "ottavo": 8, "nono": 9, "decimo": 10
        }
        
        # Combina step e risultati
        i = 1
        while i < len(step_parts):
            if i + 1 < len(step_parts):
                scenario_name = step_parts[i].strip().lower()
                scenario_steps = step_parts[i + 1].strip()
                
                # Trova risultati corrispondenti
                scenario_results = ""
                try:
                    result_idx = result_parts.index(scenario_name)
                    if result_idx + 1 < len(result_parts):
                        scenario_results = result_parts[result_idx + 1].strip()
                except ValueError:
                    pass
                
                scenario_num = scenario_map.get(scenario_name, i // 2 + 1)
                
                # Parse steps in lista
                steps_list = self._parse_bullet_list(scenario_steps)
                
                scenarios.append({
                    "id": f"scenario_{scenario_num}",
                    "name": f"Scenario {scenario_num} (extracted)",
                    "execution_steps": steps_list,
                    "expected_results": [scenario_results] if scenario_results else [],
                    "prompt_hints": None
                })
            i += 2
        
        return scenarios
    
    def _parse_bullet_list(self, text: str) -> list[str]:
        """Converte testo con bullet in lista di stringhe."""
        lines = text.split('\n')
        items = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            # Rimuovi bullet markers
            line = re.sub(r'^[•\-\*]\s*', '', line)
            if line:
                items.append(line)
        return items


def parse_test_document(file_path: str) -> Dict[str, str]:
    """
    Funzione di convenienza per parsare un documento.
    
    Args:
        file_path: Percorso al file da parsare
    
    Returns:
        Dict con sezioni estratte
    """
    parser = TestDocumentParser(file_path)
    return parser.parse()


# === CLI per testing ===
if __name__ == "__main__":
    import sys
    import json
    
    if len(sys.argv) < 2:
        print("Uso: python document_parser.py <file_path>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    try:
        result = parse_test_document(file_path)
        print(json.dumps(result, indent=2, ensure_ascii=False))
    except Exception as e:
        print(f"Errore: {e}")
        sys.exit(1)
